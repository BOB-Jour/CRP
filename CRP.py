import requests
from bs4 import BeautifulSoup
import argparse
import datetime
import re
import openpyxl # for xlsx
import os
import json

'''
# 필수 설치
$ pip3 install beautifulsoup4
$ pip3 install openpyxl
    - 참고 : https://book.coalastudy.com/data-crawling/week-5/stage-2
'''

DEBUG = False
def DEBUG_print(msg):
    if(DEBUG):
        print("[DEBUG] "+msg)

def DEBUG_print_result(result):
    if(DEBUG):
        print(result, type(result)) # result  

class CRP():
    def __init__(self, update_max_datetime=None, update_min_datetime=None, max_result=None, cumulative_flag=True, cumulative_file_path='./CRP_Data.xlsx'): 
        '''
        cumulative  : 누적 여부 결정
        cumulative_file_path : 누적 데이터를 가지고 있는 파일 경로 : 기본값 : './CRP_Data.xlsx' # 다른걸로 바꿔서 사용하다 에러가 날 수 있으니 가능하면 쓰지 말 것 # 중복 체크 지원

        [url params]
        updated-max : 검색 마지막 날짜 : 없으면 가장 최신 업데이트까지 # format : 2021-10-01
        updated-min : 검색 시작 날짜 : 없으면 가장 최초 업데이트까지   # format : 2021-09-01
        max-results : 한번에 검색할 결과 개수 (전부 다 한페이지에 띄워주지는 않음)
        '''
        self.url = 'https://chromereleases.googleblog.com/' 
        self.url_prefix = 'https://chromereleases.googleblog.com/search/label/Stable%20updates?'
        self.get_url = self.url_prefix
        if(update_max_datetime != None):
            self.get_url += '&updated-max={update_max_datetime}'.format(update_max_datetime=update_max_datetime+"T23:59:59-00:00")
        if(update_min_datetime != None):
            self.get_url += '&updated-min={update_min_datetime}'.format(update_min_datetime=update_min_datetime+"T00:00:00-00:00")
        if(max_result != None):
            self.get_url += '&max-results={max_result}'.format(max_result=max_result)

        # 종합 + 빈도
        self.all = [] # 종합 데이터 # list
        self.reward_count = {'All':0} # 금액 빈도 # dict + list 
        self.high_medi_low_count = {'All':0,'High':[],'Medium':[],'Low':[]} # 취약점 심각도 빈도 # dict + list 
        self.vulntype_count = {'All':0} # 취약점 빈도 # dict + list 
        self.vulnvector_count = {'All':0} # 취약점 터진곳 빈도 # dict + list 
        self.who_count = {'All':0} # 찾은 사람 빈도 # dict + list 
        self.date_all = {} # 날짜 # dict + list ## {"날짜1":[종합, ...], "날짜2":[종합, ...], ...}
        self.date = [] # date_all을 역순으로 만들때 쓸 것

        # load 할때 저장할 이전 값
        self.min_date = None
        self.max_date = None

        self.wb = openpyxl.Workbook()

        if(cumulative_flag):
            self.wb = self.load(cumulative_file_path) 
            self.show() # test

    def get_data(self, url):
        print(url)
        response = requests.get(url)

        # div로 구성되어있음
        if response.status_code == 200:
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')
            div_post = soup.select("div.post")
            for i in div_post:
                text_data = [v for v in i.text.split('\n') if v]
                if (text_data[0] != "Stable Channel Update for Desktop"):
                    continue
                print("title   :",text_data[0])
                print("date    :",text_data[1])
                
                if(text_data[1] in self.wb.sheetnames): # 중복 체크
                    continue

                # 종합
                print("") 
                p = re.compile('\[\$\w+\]\[\d{7,8}\][- ]?\w+[- ]?CVE-\d+-\d+[- ]?\:[- ]?\D+[- ]?in[- ]?\D+\.[- ]?Reported[- ]?by[- ]?.+?[- ]?on[- ]?\d+-\d+-\d+') 
                vuln_list = p.findall(text_data[2])
                self.date_all[text_data[1]] = vuln_list
                self.date.append(text_data[1])

                for i in vuln_list:
                    print(i) # test
                    self.all.append(i)

                    # crbug 이슈번호 추출 정규식
                    p = re.compile('\d{7,8}') 
                    issue_num = p.findall(i)[0]
                    print(issue_num)

                    # CVE 추출 정규식
                    p = re.compile('CVE-\d+-\d+') 
                    print(p.findall(i))

                    # 리워드, 심각도, 취약점 종류, 취약점 벡터, 찾은 사람 및 소속 - 추출 정규식들 : 코드 최적화
                    _regex = [re.compile('\$\w+'), re.compile('High|Medium|Low|Critical'), re.compile('\: (\D+) in') , re.compile('in (\D+)\.') , re.compile('Reported by (.+?) on')] 
                    _counts = [self.reward_count, self.high_medi_low_count, self.vulntype_count, self.vulnvector_count, self.who_count]
                    for idx, regex in enumerate(_regex):
                        print(regex.findall(i))
                        try:
                            _counts[idx][regex.findall(i)[0]].append(issue_num)
                        except KeyError:
                            _counts[idx][regex.findall(i)[0]] = [issue_num]
                        _counts[idx]['All'] += 1
                    
                    # 날짜 추출 정규식
                    p = re.compile('on (\d+-\d+-\d+)') 
                    print(p.findall(i))

                    print("")

        next_page_url = soup.find_all("a","blog-pager-older-link") # a태그의 class="blog-pager-older-link"
        DEBUG_print(next_page_url)
        if(len(next_page_url)):
            DEBUG_print("next page url :", next_page_url[0]['href']) 
            self.get_data(next_page_url[0]['href'])

    def run(self, file_path=None):
        self.get_data(self.get_url)
        try:
            self.save(file_path)
        except PermissionError:
            print("[!] PermissionError : File Open...")

    def save(self, file_path=None): # to xlsx
        global args

        try:
            intro_sheet = self.wb['Introduce'] # 뭐넣지..?
        except KeyError: # 최초 생성시
            intro_sheet = self.wb.active # 활성화 된 시트1 선택  # 종합 시트
            intro_sheet.title = "Introduce"
        try:
            reward_sheet = self.wb['Reward']
        except KeyError:
            reward_sheet = self.wb.create_sheet('Reward') # 시트 생성 # 리워드
        try:
            severity_sheet = self.wb['Severity']
        except KeyError:
            severity_sheet = self.wb.create_sheet('Severity') # 시트 생성 # 심각도
        try:
            vuln_sheet = self.wb['Vuln']
        except KeyError:
            vuln_sheet = self.wb.create_sheet('Vuln') # 시트 생성 # 취약점 종류
        try:
            vulnvector_sheet = self.wb['Vuln_Vector']
        except KeyError:
            vulnvector_sheet = self.wb.create_sheet('Vuln_Vector') # 시트 생성 # 취약점 터진 곳
        try:
            who_sheet = self.wb['Who']
        except KeyError:
            who_sheet = self.wb.create_sheet('Who') # 시트 생성 # 찾은 사람

        # Introduce.... 
        intro_sheet['A1'] = "Team"
        intro_sheet['B1'] = "Team BoB-jour (BoB10)"
        intro_sheet['A2'] = "Git"
        intro_sheet['B2'] = "https://github.com/BOB-Jour/CRP-Chrome_Releases_Parser-"
        
        intro_sheet['A3'] = "수집 시작 날짜"
        if (args.updated_min != None) and (self.min_date != None):
            if (datetime.datetime.strptime(args.updated_min, "%Y-%m-%d") < self.min_date) : # 인자로 받은 값이 더 오래된 값인지 체크
                intro_sheet['B3'] = str(args.updated_min)
            else:
                intro_sheet['B3'] = self.min_date.strftime("%Y-%m-%d")
        elif(args.load_path == 'False'): # load를 하지 않는 경우
            intro_sheet['B3'] = str(args.updated_min)
        else : # None인 경우가 더 과거의 데이터를 포함한다는 뜻이기 때문에 그대로 None으로 저장 # 다만 load를 하지 않은 경우에도 같은 None임...
            intro_sheet['B3'] = str(None)

        intro_sheet['A4'] = "수집 마지막 날짜"
        if (args.updated_max != None) and (self.max_date != None):
            if (datetime.datetime.strptime(args.updated_max, "%Y-%m-%d") > self.max_date) : # 인자로 받은 값이 더 오래된 값인지 체크
                intro_sheet['B4'] = str(args.updated_max)
            else:
                intro_sheet['B4'] = self.max_date.strftime("%Y-%m-%d")
        elif(args.load_path == 'False'): # load를 하지 않는 경우
            intro_sheet['B4'] = str(args.updated_max)
        else: # max의 경우 None이면 가장 최신날짜가 되기 떄문에 오늘의 날짜를 적는다.
            intro_sheet['B4'] = datetime.datetime.now().strftime("%Y-%m-%d")

        # 날짜별 종합 시트 # 역순으로(가장 오래된거 순서대로)
        self.date.reverse()
        for value in self.date:
            if(value in self.wb.sheetnames): # 중복 체크
                DEBUG_print("continue:",value) # test
                continue
            tmp_sheet = self.wb.create_sheet(value) 
            for i,v in enumerate(self.date_all[value]):
                tmp_sheet['A%d' % (i+1)] = v
        
        _counts = [self.reward_count, self.high_medi_low_count, self.vulntype_count, self.vulnvector_count, self.who_count]
        _sheets = [reward_sheet, severity_sheet, vuln_sheet, vulnvector_sheet, who_sheet]
        
        # 빈도 시트 
        for counts_i,counts_v in enumerate(_counts):
            count = 1 
            for key, value in counts_v.items():
                _sheets[counts_i]['A%d' % count] = key
                if(key == 'All'):
                    _sheets[counts_i]['B%d' % count] = value
                    count += 1
                    continue
                _sheets[counts_i]['B%d' % count] = len(value)
                _sheets[counts_i]['C%d' % count] = json.dumps(value)
                count += 1

        if(file_path != None):
            self.wb.save(file_path)
        else:
            self.wb.save('./CRP_Data.xlsx') # default

    def load(self, file_path): # xlsx 파일 로드
        if(not os.path.isfile(file_path)):
            return openpyxl.Workbook()
        wb = openpyxl.load_workbook(file_path)
        reward_sheet = wb['Reward']
        severity_sheet = wb['Severity']
        vuln_sheet = wb['Vuln']
        vulnvector_sheet = wb['Vuln_Vector']
        who_sheet = wb['Who'] 
        DEBUG_print(reward_sheet._cells[(1,1)].value) # test 
        DEBUG_print(reward_sheet._cells[(1,2)].value) # test
        
        introduce_sheet = wb['Introduce']
        if (introduce_sheet._cells[(3,2)].value != 'None'):
            self.min_date = datetime.datetime.strptime(introduce_sheet._cells[(3,2)].value, "%Y-%m-%d")
        if (introduce_sheet._cells[(4,2)].value != 'None'):
            self.max_date = datetime.datetime.strptime(introduce_sheet._cells[(4,2)].value, "%Y-%m-%d")
        _counts = [self.reward_count, self.high_medi_low_count, self.vulntype_count, self.vulnvector_count, self.who_count]
        _sheets = [reward_sheet, severity_sheet, vuln_sheet, vulnvector_sheet, who_sheet]

        for sheets_i,sheets_v in enumerate(_sheets): # 값 로드
            count = 1 
            for key, value in sheets_v._cells.items(): # (1,1), (1,2), (2,1), (2,2)...
                if(count == 1): # All
                    _counts[sheets_i][value.value] = sheets_v._cells[(count,2)].value
                    count += 1
                    continue
                elif(count == key[0]):
                    _counts[sheets_i][value.value] = json.loads(sheets_v._cells[(count,3)].value)
                    count += 1
        return wb

    def show(self):
        global DEBUG
        if(DEBUG):
            print("=== Show ===")
            print("reward_count         : ", self.reward_count)
            print("high_medi_low_count : ", self.high_medi_low_count)
            print("vulntype_count      : ", self.vulntype_count)
            print("vulnvector_count    : ", self.vulnvector_count)
            print("who_count           : ", self.who_count)

    def push_slack(self):
        # 미구현
        pass

def argparse_init(): 
    parser = argparse.ArgumentParser(description='CRP(Chrome_Releases_Parser)')
    parser.add_argument('--updated_max', '-M', help='updated-max : 검색 마지막 날짜 : 없으면 가장 최신 업데이트까지 # format : 2021-10-01', default=None)
    parser.add_argument('--updated_min', '-m', help='updated-min : 검색 시작 날짜 : 없으면 가장 최초 업데이트부터, 웬만하면 주는게 좋다.  # format : 2021-09-01', default=None)
    parser.add_argument('--max_results', '-r', help='max-results : 한번에 검색할 결과 개수', default=None)
    parser.add_argument('--load_path', '-l', help='load cumulative file path : 누적(로드)시킬 파일 경로. 옵션 안주면 ./CRP_Data.xlsx에 자동으로 누적. False 입력시 누적 X', default='./CRP_Data.xlsx')
    parser.add_argument('--save_path', '-s', help='save file path : 저장할 파일 경로. 옵션 안주면 ./CRP_Data.xlsx 이름으로 저장', default=None)
    
    return parser

if __name__ == "__main__":

    parser = argparse_init()
    args = parser.parse_args()
    if(args.updated_max == None) and (args.updated_min == None):
        if(input("[Warning] 정말로 최초 날짜부터 최근 데이터까지 전부 수집하시겠습니까? [Y/n] ") != 'Y'):
            parser.print_help()
            os._exit(0)
    if(args.load_path == "False"):
        crp = CRP(update_max_datetime=args.updated_max,update_min_datetime=args.updated_min,max_result=args.max_results,cumulative_flag=False)
    else:
        crp = CRP(update_max_datetime=args.updated_max,update_min_datetime=args.updated_min,max_result=args.max_results,cumulative_flag=True,cumulative_file_path=args.load_path)
    crp.run(args.save_path)
    crp.show() # test